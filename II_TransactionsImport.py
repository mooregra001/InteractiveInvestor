import pandas as pd
import os
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import copy
import json
from II_Constants import (
    TRANSACTIONS_CSV_PATH, DEFAULT_EXCEL_PATH, TRANSACTIONS_SHEET,
    TRANSACTIONS_MAX_COLUMN, TRANSACTIONS_FORMULA
)
from II_Config import load_config  # Import load_config from II_Config.py

# File paths and constants
config = load_config()
excel_path = config.get("excel_path", DEFAULT_EXCEL_PATH)  # Use config.json with fallback
csv_path = TRANSACTIONS_CSV_PATH

def populate_formulas(sheet, start_row, end_row):
    """Populate column A with the formula from start_row to end_row."""
    formula = TRANSACTIONS_FORMULA
    for row in range(start_row, end_row + 1):
        sheet[f'A{row}'] = formula.format(row=row)
    print(f"Formula populated in column A from row {start_row} to {end_row}")

def copy_row_formatting(sheet, source_row, start_row, end_row):
    """Copy number_format and fill from source_row to rows from start_row to end_row for columns A:M."""
    for row in range(start_row, end_row + 1):
        for col in range(1, TRANSACTIONS_MAX_COLUMN + 1):
            source_cell = sheet.cell(row=source_row, column=col)
            target_cell = sheet.cell(row=row, column=col)
            target_cell.number_format = source_cell.number_format
            target_cell.fill = copy.copy(source_cell.fill) if source_cell.fill else PatternFill()
    print(f"Number formats and fills copied from row {source_row} to rows {start_row} to {end_row} for columns A to {get_column_letter(TRANSACTIONS_MAX_COLUMN)}")

try:
    # Verify CSV file exists
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found at: {csv_path}")
    
    # Verify Excel file exists
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")

    # Read CSV file with UTF-8-SIG encoding to handle BOM
    print("Reading CSV file...")
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    print(f"CSV rows read: {len(df)}")

    # Clean column names by removing BOM characters
    df.columns = df.columns.str.replace('\ufeff', '')
    print("Column names in CSV:", list(df.columns))

    # Replace 'n/a' with empty string
    df = df.replace('n/a', '')

    # Clean numeric columns by removing currency symbols (£) and commas
    for col in ['Quantity', 'Price', 'Debit', 'Credit', 'Running Balance']:
        if col in df.columns:
            df[col] = df[col].replace('[\£,]', '', regex=True).replace('', None).astype(float, errors='ignore')

    # Sort by Date in ascending order
    print("Sorting data by Date...")
    if 'Date' not in df.columns:
        raise ValueError("Column 'Date' not found in CSV. Available columns: " + str(list(df.columns)))
    
    try:
        # Convert Date to datetime
        df['Date'] = pd.to_datetime(df['Date'], dayfirst=True, errors='coerce')
        if df['Date'].isna().any():
            print("Warning: Some dates in 'Date' could not be parsed. Invalid dates will be left as-is:")
            print(df[df['Date'].isna()][['Date']])
        
        # Convert Settlement Date to datetime
        if 'Settlement Date' in df.columns:
            df['Settlement Date'] = pd.to_datetime(df['Settlement Date'], dayfirst=True, errors='coerce')
            if df['Settlement Date'].isna().any():
                print("Warning: Some dates in 'Settlement Date' could not be parsed. Invalid dates will be left as-is:")
                print(df[df['Settlement Date'].isna()][['Settlement Date']])
        else:
            print("Warning: 'Settlement Date' column not found in CSV. Available columns: " + str(list(df.columns)))

        df = df.sort_values(by='Date', ascending=True, na_position='last')
    except Exception as e:
        print(f"Warning: Failed to sort by Date or parse dates due to: {str(e)}. Proceeding without sorting.")

    # Apply negative sign to Quantity when Credit has a value, unless Description starts with "Div" or "GROSS INTEREST"
    df['Quantity'] = df.apply(
        lambda row: -float(row['Quantity']) if pd.notnull(row['Credit']) and row['Credit'] != '' 
        and not (str(row['Description']).startswith('Div') or str(row['Description']).startswith('GROSS INTEREST')) 
        else float(row['Quantity']) if pd.notnull(row['Quantity']) else None,
        axis=1
    )

    # Apply negative sign to Debit when Debit has a value
    df['Debit'] = df.apply(
        lambda row: -float(row['Debit']) if pd.notnull(row['Debit']) else None,
        axis=1
    )

    # Ensure Credit and Running Balance are float
    df['Credit'] = df['Credit'].astype(float, errors='ignore')
    df['Running Balance'] = df['Running Balance'].astype(float, errors='ignore')

    # Load existing Excel file
    print("Loading Excel file...")
    book = openpyxl.load_workbook(excel_path)
    
    # Verify Transactions sheet exists
    if TRANSACTIONS_SHEET not in book.sheetnames:
        raise ValueError("Transactions sheet not found in Excel file")

    # Get the Transactions sheet
    sheet = book[TRANSACTIONS_SHEET]

    # Find last non-empty row based on column B (Date)
    last_row = 1
    for row in range(1, sheet.max_row + 1):
        cell = sheet[f'B{row}']
        if cell.value is None or (isinstance(cell.value, str) and not cell.value.strip()):
            last_row = max(1, row - 1)
            break
        last_row = row
    print(f"Last non-empty row in column B: {last_row}")

    # Use pandas to append data to Excel, starting at column B (index 1) and the next available row
    with pd.ExcelWriter(excel_path, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
        writer.workbook = book
        df.to_excel(
            writer,
            sheet_name=TRANSACTIONS_SHEET,
            startrow=last_row,  # Start at the row after the last non-empty row
            startcol=1,  # Start at column B (index 1)
            index=False,  # Don't write DataFrame index
            header=False  # Don't write column headers
        )

    print(f"Data successfully written to {excel_path} in Transactions sheet")

    # Re-load workbook for formatting
    book = openpyxl.load_workbook(excel_path)
    sheet = book[TRANSACTIONS_SHEET]

    # Find the last row with data in column B
    last_row = sheet.max_row
    for row in range(last_row, 1, -1):
        if sheet[f'B{row}'].value is not None:
            last_row = row
            break

    # Find the first empty cell in column A
    first_empty_row = 2
    for row in range(2, last_row + 1):
        if sheet[f'A{row}'].value is None:
            first_empty_row = row
            break

    # Populate formulas
    populate_formulas(sheet, first_empty_row, last_row)

    # Copy formatting from the row above first_empty_row
    if first_empty_row > 1:
        copy_row_formatting(sheet, first_empty_row - 1, first_empty_row, last_row)
    else:
        print("Error: No row above first_empty_row to copy formatting from")

    # Save the workbook
    book.save(excel_path)
    print("File saved successfully")

except FileNotFoundError as e:
    print(f"Error: {str(e)}")
except PermissionError:
    print(f"Error: Permission denied when accessing {excel_path}. Ensure the file is not open in another application.")
except ValueError as ve:
    print(f"Error: {ve}")
except Exception as e:
    print(f"An unexpected error occurred: {str(e)}")