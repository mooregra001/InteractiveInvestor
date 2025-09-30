import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import PatternFill
import os
import copy
import json
from II_Constants import (
    DEFAULT_EXCEL_PATH, INVESTMENTS_CSV_PATH, INVESTMENTS_SHEET,
    INVESTMENTS_START_CELL, INVESTMENTS_MAX_COLUMN, INVESTMENTS_FORMULA
)
from II_Config import load_config  # Import load_config from II_Config.py

# File paths and constants
config = load_config()
excel_path = config.get("excel_path", DEFAULT_EXCEL_PATH)  # Use config.json with fallback
csv_path = INVESTMENTS_CSV_PATH
sheet_name = INVESTMENTS_SHEET
start_cell = INVESTMENTS_START_CELL

# Rest of the script remains unchanged
...
def clear_excel_range(file_path, sheet_name, start_cell):
    """
    Clears the content of the Excel sheet starting from the specified cell
    (B2) to the extent of the current region, preserving headers and formatting.
    """
    if not os.path.exists(file_path):
        raise FileNotFoundError(f"Excel file not found at: {file_path}")
    
    wb = load_workbook(file_path)
    ws = wb[sheet_name]
    
    start_col = ws[start_cell].column
    start_row = 2
    
    max_row = ws.max_row
    max_col = ws.max_column
    
    for row in range(start_row, max_row + 1):
        for col in range(start_col, max_col + 1):
            ws.cell(row=row, column=col).value = None  # Clear value only, preserve formatting
    
    wb.save(file_path)
    wb.close()
    print(f"Successfully cleared the data range starting from {start_cell} in sheet '{sheet_name}' of {file_path}, preserving headers and formatting.")

def clean_price_or_avg_price(value, is_first_row=False):
    """
    Custom cleaning for 'Price' column.
    - First row: Remove $ if present.
    - Other rows: Remove 'p' suffix if numeric.
    - Preserve non-numeric strings.
    """
    if pd.isna(value):
        return value
    value_str = str(value).strip()
    
    if not value_str.replace('$', '').replace('p', '').replace('-', '').replace('.', '').replace(',', '').isdigit():
        return value_str
    
    if is_first_row:
        value_str = value_str.replace('$', '')
    else:
        value_str = value_str.replace('p', '')
    
    value_str = value_str.replace(',', '')
    try:
        return float(value_str)
    except ValueError:
        return value_str

def clean_percentage(value):
    """
    Clean percentage columns ('Day Gain/Loss %', 'Gain/Loss %').
    Remove % suffix and convert to float, handling - prefix.
    """
    if pd.isna(value):
        return value
    value_str = str(value).strip()
    value_str = value_str.replace('%', '').replace(',', '')
    try:
        return float(value_str)
    except ValueError:
        return value_str

def populate_formulas(sheet, start_row, end_row, source_row=None):
    """
    Populate column A with the formula from start_row to end_row, copying formatting
    from source_row (defaults to start_row - 1 or row 2 if not specified).
    """
    formula = INVESTMENTS_FORMULA
    # Determine source row for formatting (use row above start_row or row 2)
    source_row = source_row or (start_row - 1 if start_row > 2 else 2)
    source_cell = sheet[f'A{source_row}']
    
    for row in range(start_row, end_row + 1):
        cell = sheet[f'A{row}']
        cell.value = formula.format(row=row)
        # Copy formatting from source_cell
        cell.number_format = source_cell.number_format
        cell.fill = copy.copy(source_cell.fill) if source_cell.fill else PatternFill()
            
    print(f"Formula populated in column A from row {start_row} to {end_row}, "
          f"with formatting copied from A{source_row}")

def import_csv_to_excel(csv_path, excel_path, sheet_name, start_cell):
    """
    Imports data from a CSV file to an Excel sheet starting at the specified cell (B2),
    applies formulas with copied formatting in column A, preserves existing formatting elsewhere,
    and saves the workbook.
    """
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found at: {csv_path}")
    
    # Load workbook
    try:
        wb = openpyxl.load_workbook(excel_path)
    except FileNotFoundError:
        raise FileNotFoundError(f"Excel file not found at: {excel_path}")
    
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Sheet '{sheet_name}' not found in the workbook")
    
    ws = wb[sheet_name]
    
    # Clear existing data range, preserving formatting
    clear_excel_range(excel_path, sheet_name, start_cell)
    
    # Reload workbook after clearing
    wb = openpyxl.load_workbook(excel_path)
    ws = wb[sheet_name]
    
    # Read and clean CSV data
    print("Reading CSV file...")
    df = pd.read_csv(csv_path, encoding='utf-8-sig')
    print(f"CSV rows read: {len(df)}")
    
    df.columns = df.columns.str.replace('\ufeff', '')
    print("Column names in CSV:", list(df.columns))
    
    df = df.replace('n/a', '')
    
    # Define numeric columns for cleaning (but not for formatting)
    numeric_columns = [
        'Price', 'Day Gain/Loss', 'Day Gain/Loss %', 'Market Value £',
        'Book Cost', 'Gain/Loss', 'Gain/Loss %'
    ]
    numeric_columns = [col for col in numeric_columns if col in df.columns]
    print("Processing numeric columns:", numeric_columns)
    
    # Clean numeric columns
    for col in numeric_columns:
        if col == 'Price':
            df[col] = [clean_price_or_avg_price(val, is_first_row=(idx == 0)) for idx, val in enumerate(df[col])]
        elif col in ['Day Gain/Loss %', 'Gain/Loss %']:
            df[col] = df[col].apply(clean_percentage)
        else:
            df[col] = df[col].replace('[\£\$,]', '', regex=True).replace('', None)
            df[col] = pd.to_numeric(df[col], errors='coerce').astype(float, errors='ignore')
    
    # Write DataFrame to Excel, preserving existing formatting
    start_col = ws[start_cell].column
    start_row = ws[start_cell].row
    
    for r_idx, row in enumerate(df.itertuples(index=False), start=start_row):
        for c_idx, value in enumerate(row, start=start_col):
            cell = ws.cell(row=r_idx, column=c_idx)
            cell.value = value  # Write value only, preserve existing formatting
    
    # Find the last row with data in the imported data
    last_row = ws.max_row
    for row in range(last_row, 1, -1):
        if ws[f'B{row}'].value is not None:
            last_row = row
            break
    
    # Find the first empty cell in column A
    first_empty_row = 2
    for row in range(2, last_row + 1):
        if ws[f'A{row}'].value is None:
            first_empty_row = row
            break
    
    # Populate formulas in column A, copying formatting from the row above or row 2
    populate_formulas(ws, first_empty_row, last_row)
    
    # Save the workbook
    try:
        wb.save(excel_path)
        wb.close()
        print(f"Successfully imported data from {csv_path} to sheet '{sheet_name}' starting at {start_cell} in {excel_path}, "
              f"with column A formatting copied from A{first_empty_row - 1 if first_empty_row > 2 else 2}.")
    except PermissionError:
        raise PermissionError(f"Permission denied when saving {excel_path}. Ensure the file is not open in another application.")

# Execute the import
try:
    import_csv_to_excel(csv_path, excel_path, sheet_name, start_cell='B2')
except FileNotFoundError as e:
    print(f"Error: {e}")
except PermissionError as e:
    print(f"Error: {e}")
except ValueError as e:
    print(f"Error: {e}")
except Exception as e:
    print(f"An unexpected error occurred: {str(e)}")
    import sys
    sys.exit(1)